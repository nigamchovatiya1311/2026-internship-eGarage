from .decorators import role_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Avg, Count
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db.models import Avg as AvgF
from django.core.mail import send_mail
from django.conf import settings
import csv
from django.contrib.auth import logout as auth_logout
from django.contrib.auth import logout
from datetime import date, timedelta
from calendar import month_abbr
import uuid
from .models import Payments

# ── Import your models ──────────────────────────────────────
from core.models import User
from .models import (
    ServiceProvider,
    CustomerProfile,
    Vehicle,
    Services,
    Bookings,
    Payments,
    Invoice,
    Review,
    Notification,
)


# ============================================================
#  ADMIN REDIRECT
# ============================================================
@role_required(allowed_roles=["admin"])
def admindashboard(request):
    return redirect('admin_overview')


# ============================================================
#  SHARED BASE CONTEXT  (sidebar badges — live counts)
# ============================================================
def get_base_context(request):
    return {
        'total_users':          User.objects.count(),
        'pending_providers':    ServiceProvider.objects.filter(approvalStatus='pending').count(),
        'new_bookings':         Bookings.objects.filter(bookingStatus='pending').count(),
        'unread_notifications': Notification.objects.filter(isRead=False).count(),
        'open_disputes':        0,
        'notifications_list':   Notification.objects.filter(isRead=False).order_by('-createdAt')[:5],
    }


# ============================================================
#  1. OVERVIEW  —  /admin-panel/
# ============================================================
@role_required(allowed_roles=["admin"])
def overview(request):
    today = date.today()

    total_users     = User.objects.count()
    total_providers = ServiceProvider.objects.count()
    total_bookings  = Bookings.objects.count()
    monthly_revenue = (
        Payments.objects
        .filter(paymentStatus='completed', paymentDate__month=today.month, paymentDate__year=today.year)
        .aggregate(total=Sum('amount'))['total'] or 0
    )

    # ── Last 6 months revenue chart ────────────────────────────
    revenue_by_month = (
        Payments.objects
        .filter(paymentStatus='completed')
        .annotate(month=TruncMonth('paymentDate'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    # Build a dict keyed by (year, month)
    rev_map = {}
    for row in revenue_by_month:
        if row['month']:
            rev_map[(row['month'].year, row['month'].month)] = float(row['total'] or 0)

    # Generate last 6 months including current
    monthly_revenue_chart = []
    for i in range(5, -1, -1):
        d     = today.replace(day=1) - timedelta(days=1) * 0
        month = (today.month - i - 1) % 12 + 1
        year  = today.year + ((today.month - i - 1) // 12)
        amt   = rev_map.get((year, month), 0)
        monthly_revenue_chart.append({
            'month':  month_abbr[month],
            'amount': int(amt),
        })

    # Normalise bar heights (max = 100%)
    max_amt = max((e['amount'] for e in monthly_revenue_chart), default=1) or 1
    for e in monthly_revenue_chart:
        e['height'] = round((e['amount'] / max_amt) * 90 + 10)  # min 10% so bar is always visible

    status_counts = Bookings.objects.aggregate(
        completed   = Count('bookingId', filter=Q(bookingStatus='completed')),
        in_progress = Count('bookingId', filter=Q(bookingStatus='in_progress')),
        pending     = Count('bookingId', filter=Q(bookingStatus='pending')),
        cancelled   = Count('bookingId', filter=Q(bookingStatus='cancelled')),
    )

    recent_bookings = (
        Bookings.objects
        .select_related('customer__user', 'provider').prefetch_related('services')
        .order_by('-createdAt')[:5]
    )

    enriched_bookings = []
    for b in recent_bookings:
        amount = 0
        if hasattr(b, 'payment'):
            amount = b.payment.amount
        enriched_bookings.append({
            'id':       b.bookingId,
            'customer': b.customer.user,
            'service':  b.service_names,
            'provider': b.provider,
            'date':     b.bookingDate,
            'amount':   amount,
            'status':   b.bookingStatus,
        })

    context = {
        **get_base_context(request),
        'active_section':       'overview',
        'today':                today.strftime('%d %b %Y'),
        'total_users':          total_users,
        'total_bookings':       total_bookings,
        'total_providers':      total_providers,
        'monthly_revenue':        monthly_revenue,
        'monthly_revenue_chart':  monthly_revenue_chart,
        'completed_bookings':   status_counts['completed'],
        'inprogress_bookings':  status_counts['in_progress'],
        'pending_bookings':     status_counts['pending'],
        'cancelled_bookings':   status_counts['cancelled'],
        'recent_bookings':      enriched_bookings,
    }
    return render(request, 'garage/Admin/overview.html', context)


# ============================================================
#  2. MANAGE USERS
# ============================================================
@role_required(allowed_roles=["admin"])
def manage_users(request):
    q      = request.GET.get('q', '').strip()
    role   = request.GET.get('role', 'all')
    status = request.GET.get('status', '')
    page   = request.GET.get('page', 1)

    qs = User.objects.order_by('-created_at')

    if q:
        qs = qs.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q)  |
            Q(email__icontains=q)
        )
    if role not in ('', 'all'):
        qs = qs.filter(role=role)
    if status == 'active':
        qs = qs.filter(is_active=True)
    elif status == 'blocked':
        qs = qs.filter(is_active=False)

    COLORS = ['#e8560a', '#1e3a5f', '#16a34a', '#7c3aed', '#f9a825', '#dc2626']
    users_with_color = []
    for i, u in enumerate(qs):
        u.avatar_color = COLORS[i % len(COLORS)]
        u.status       = 'active' if u.is_active else 'blocked'
        users_with_color.append(u)

    paginator  = Paginator(users_with_color, 20)
    users_page = paginator.get_page(page)

    context = {
        **get_base_context(request),
        'active_section':  'users',
        'users':           users_page,
        'total_users':     User.objects.count(),
        'total_customers': User.objects.filter(role='customer').count(),
        'total_providers': User.objects.filter(role='service_provider').count(),
        'total_admins':    User.objects.filter(is_staff=True).count(),
    }
    return render(request, 'garage/Admin/users.html', context)


@role_required(allowed_roles=["admin"])
def block_user(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.is_active = False
    user.save()
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    messages.success(request, f'{user.get_full_name()} blocked.')
    return redirect('admin_users')


@role_required(allowed_roles=["admin"])
def unblock_user(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.is_active = True
    user.save()
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    messages.success(request, f'{user.get_full_name()} unblocked.')
    return redirect('admin_users')


@role_required(allowed_roles=["admin"])
def approve_user(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.is_active = True
    user.save()
    messages.success(request, f'{user.get_full_name()} approved successfully.')
    return redirect('admin_users')


@role_required(allowed_roles=["admin"])
def add_user(request):
    if request.method == 'POST':
        messages.success(request, 'User created successfully.')
        return redirect('admin_users')
    context = {**get_base_context(request), 'active_section': 'users'}
    return render(request, 'garage/Admin/add_user.html', context)


# ============================================================
#  3. SERVICE PROVIDERS
# ============================================================
@role_required(allowed_roles=["admin"])
def service_providers(request):
    q        = request.GET.get('q', '').strip()
    approval = request.GET.get('approval', 'all')
    page     = request.GET.get('page', 1)

    qs = ServiceProvider.objects.select_related('user').order_by('-providerId')

    if q:
        qs = qs.filter(
            Q(garageName__icontains=q)       |
            Q(location__icontains=q)          |
            Q(user__first_name__icontains=q)
        )
    if approval not in ('', 'all'):
        qs = qs.filter(approvalStatus=approval)

    for p in qs:
        p.garage_name     = p.garageName
        p.city            = p.location
        p.working_hours   = f"{p.openingTime.strftime('%I:%M %p')} – {p.closingTime.strftime('%I:%M %p')}"
        p.avg_rating      = p.rating
        p.approval_status = p.approvalStatus

    paginator      = Paginator(qs, 20)
    providers_page = paginator.get_page(page)

    stats = ServiceProvider.objects.aggregate(
        approved = Count('providerId', filter=Q(approvalStatus='approved')),
        pending  = Count('providerId', filter=Q(approvalStatus='pending')),
        rejected = Count('providerId', filter=Q(approvalStatus='rejected')),
        avg_r    = Avg('rating'),
    )

    context = {
        **get_base_context(request),
        'active_section': 'providers',
        'providers':      providers_page,
        'approved_count': stats['approved'],
        'pending_count':  stats['pending'],
        'rejected_count': stats['rejected'],
        'avg_rating':     round(stats['avg_r'] or 0, 1),
    }
    return render(request, 'garage/Admin/providers.html', context)


@role_required(allowed_roles=["admin"])
def approve_provider(request, pk):
    provider = get_object_or_404(ServiceProvider, pk=pk)
    provider.approvalStatus = 'approved'
    provider.save()
    Notification.objects.create(
        user=provider.user,
        notificationType='general',
        title='Garage Approved',
        message=f'Your garage "{provider.garageName}" has been approved. You can now receive bookings.',
    )
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'garage_name': provider.garageName})
    messages.success(request, f'"{provider.garageName}" approved successfully.')
    return redirect('admin_providers')


@role_required(allowed_roles=["admin"])
def reject_provider(request, pk):
    provider = get_object_or_404(ServiceProvider, pk=pk)
    name = provider.garageName
    Notification.objects.create(
        user=provider.user,
        notificationType='general',
        title='Garage Application Rejected',
        message=f'Your garage "{name}" application has been rejected. Please contact support.',
    )
    provider.delete()
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    messages.success(request, f'"{name}" removed.')
    return redirect('admin_providers')


# ============================================================
#  4. CUSTOMER PROFILES
# ============================================================
@role_required(allowed_roles=["admin"])
def customer_profiles(request):
    q            = request.GET.get('q', '').strip()
    vehicle_type = request.GET.get('type', 'all')

    qs = CustomerProfile.objects.select_related('user').order_by('user__first_name')

    if q:
        qs = qs.filter(
            Q(user__first_name__icontains=q) |
            Q(vehicleNumber__icontains=q)    |
            Q(vehicleModel__icontains=q)
        )
    if vehicle_type not in ('', 'all'):
        qs = qs.filter(vehicleType=vehicle_type)

    for cp in qs:
        cp.vehicle_type   = cp.vehicleType
        cp.vehicle_number = cp.vehicleNumber
        cp.vehicle_model  = cp.vehicleModel
        cp.vehicle_year   = cp.vehicleYear
        cp.vehicle_color  = cp.vehicleColor

    context = {
        **get_base_context(request),
        'active_section':    'customers',
        'customer_profiles': qs,
    }
    return render(request, 'garage/Admin/customers.html', context)


# ============================================================
#  5. SERVICES
# ============================================================
@role_required(allowed_roles=["admin"])
def services(request):
    q        = request.GET.get('q', '').strip()
    provider = request.GET.get('provider', '')
    status   = request.GET.get('status', '')

    qs = Services.objects.select_related('providerId').order_by('serviceName')

    if q:
        qs = qs.filter(
            Q(serviceName__icontains=q) |
            Q(serviceDescription__icontains=q)
        )
    if provider:
        qs = qs.filter(providerId__garageName__icontains=provider)
    if status:
        qs = qs.filter(isAvailable=(status == 'available'))

    for svc in qs:
        svc.name             = svc.serviceName
        svc.description      = svc.serviceDescription
        svc.price            = svc.servicePrice
        svc.duration_minutes = svc.estimatedDuration
        svc.is_available     = svc.isAvailable
        svc.provider         = svc.providerId
        svc.provider.garage_name = svc.providerId.garageName

    provider_list = ServiceProvider.objects.filter(approvalStatus='approved')
    for p in provider_list:
        p.garage_name = p.garageName

    context = {
        **get_base_context(request),
        'active_section': 'services',
        'services':       qs,
        'provider_list':  provider_list,
    }
    return render(request, 'garage/Admin/services.html', context)


@role_required(allowed_roles=["admin"])
@require_POST
def save_service(request):
    service_id  = request.POST.get('service_id', '').strip()
    name        = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip()
    price       = request.POST.get('price', 0)
    duration    = request.POST.get('duration', 0)
    # ✅ FIX: modal sends provider_id (PK integer), not provider name
    provider_id = request.POST.get('provider_id', '').strip()

    if not name:
        messages.warning(request, 'Service name is required.')
        return redirect('admin_services')

    vehicle_type = request.POST.get('vehicle_type', 'all').strip()

    if service_id:
        svc = get_object_or_404(Services, pk=service_id)
        svc.serviceName        = name
        svc.serviceDescription = description
        svc.servicePrice       = price
        svc.estimatedDuration  = duration
        svc.vehicleType        = vehicle_type
        # ✅ FIX: update provider if a new one was selected
        if provider_id:
            provider_obj = get_object_or_404(ServiceProvider, pk=provider_id)
            svc.providerId = provider_obj
        svc.save()
        messages.success(request, f'"{name}" updated successfully.')
    else:
        # ✅ FIX: lookup by PK, not garageName — raises 400 if missing
        if not provider_id:
            messages.error(request, 'Please select a provider.')
            return redirect('admin_services')
        provider_obj = get_object_or_404(ServiceProvider, pk=provider_id)
        Services.objects.create(
            serviceName        = name,
            serviceDescription = description,
            servicePrice       = price,
            estimatedDuration  = duration,
            vehicleType        = vehicle_type,
            providerId         = provider_obj,
        )
        messages.success(request, f'"{name}" added successfully.')

    return redirect('admin_services')


# ============================================================
#  6. MONITOR BOOKINGS
# ============================================================
@role_required(allowed_roles=["admin"])
def monitor_bookings(request):
    q      = request.GET.get('q', '').strip()
    status = request.GET.get('status', 'all')
    bdate  = request.GET.get('date', '')
    page   = request.GET.get('page', 1)

    qs = (
        Bookings.objects
        .select_related('customer__user', 'provider').prefetch_related('services')
        .order_by('-bookingDate', '-bookingTime')
    )

    if q:
        qs = qs.filter(
            Q(bookingId__icontains=q)                  |
            Q(customer__user__first_name__icontains=q) |
            Q(services__serviceName__icontains=q)       |
            Q(provider__garageName__icontains=q)
        )
    if status not in ('', 'all'):
        qs = qs.filter(bookingStatus=status)
    if bdate:
        qs = qs.filter(bookingDate=bdate)

    for b in qs:
        b.id        = b.bookingId
        b.status    = b.bookingStatus
        b.date      = b.bookingDate
        b.time_slot = b.bookingTime
        b.service_display      = b.service_names
        b.provider.garage_name = b.provider.garageName
        b.amount = b.payment.amount if hasattr(b, 'payment') and b.payment else 0
        b.vehicle_number = (
            b.customer.vehicleNumber if hasattr(b.customer, 'vehicleNumber') else '—'
        )

    counts = Bookings.objects.aggregate(
        completed   = Count('bookingId', filter=Q(bookingStatus='completed')),
        in_progress = Count('bookingId', filter=Q(bookingStatus='in_progress')),
        pending     = Count('bookingId', filter=Q(bookingStatus='pending')),
        cancelled   = Count('bookingId', filter=Q(bookingStatus='cancelled')),
    )

    paginator     = Paginator(qs, 20)
    bookings_page = paginator.get_page(page)

    context = {
        **get_base_context(request),
        'active_section':   'bookings',
        'bookings':         bookings_page,
        'total_bookings':   Bookings.objects.count(),
        'completed_count':  counts['completed'],
        'inprogress_count': counts['in_progress'],
        'pending_count':    counts['pending'],
        'cancelled_count':  counts['cancelled'],
        'today_iso':        date.today().isoformat(),
    }
    return render(request, 'garage/Admin/bookings.html', context)


# ============================================================
#  7. PAYMENTS
# ============================================================
@role_required(allowed_roles=["admin"])
def payments(request):
    q      = request.GET.get('q', '').strip()
    method = request.GET.get('method', 'all')
    status = request.GET.get('status', '')
    page   = request.GET.get('page', 1)

    qs = (
        Payments.objects
        .select_related('booking__customer__user')
        .order_by('-paymentDate')
    )

    if q:
        qs = qs.filter(
            Q(transactionId__icontains=q) |
            Q(booking__customer__user__first_name__icontains=q)
        )
    if method not in ('', 'all'):
        qs = qs.filter(paymentMethod=method)
    if status:
        qs = qs.filter(paymentStatus=status)

    for p in qs:
        p.id             = p.paymentId
        p.method         = p.paymentMethod
        p.status         = p.paymentStatus
        p.transaction_id = p.transactionId
        p.created_at     = p.paymentDate

    totals = Payments.objects.aggregate(
        collected       = Sum('amount', filter=Q(paymentStatus='completed')),
        pending_amount  = Sum('amount', filter=Q(paymentStatus='pending')),
        failed_amount   = Sum('amount', filter=Q(paymentStatus='failed')),
        refunded_amount = Sum('amount', filter=Q(paymentStatus='refunded')),
    )

    paginator     = Paginator(qs, 20)
    payments_page = paginator.get_page(page)

    context = {
        **get_base_context(request),
        'active_section':  'payments',
        'payments':        payments_page,
        'collected':       totals['collected']       or 0,
        'pending_amount':  totals['pending_amount']  or 0,
        'failed_amount':   totals['failed_amount']   or 0,
        'refunded_amount': totals['refunded_amount'] or 0,
    }
    return render(request, 'garage/Admin/payments.html', context)


# ============================================================
#  8. INVOICES
# ============================================================
@role_required(allowed_roles=["admin"])
def invoices(request):
    q    = request.GET.get('q', '').strip()
    page = request.GET.get('page', 1)

    qs = (
        Invoice.objects
        .select_related('booking__customer__user')
        .order_by('-invoiceDate')
    )

    if q:
        qs = qs.filter(
            Q(invoiceNumber__icontains=q) |
            Q(booking__customer__user__first_name__icontains=q)
        )

    for inv in qs:
        inv.id              = inv.invoiceId
        inv.created_at      = inv.invoiceDate
        inv.total_amount    = inv.totalAmount
        inv.tax_amount      = inv.taxAmount
        inv.discount_amount = inv.discountAmount
        inv.service_amount  = inv.totalAmount - inv.taxAmount + inv.discountAmount

    paginator     = Paginator(qs, 20)
    invoices_page = paginator.get_page(page)

    context = {
        **get_base_context(request),
        'active_section': 'invoices',
        'invoices':       invoices_page,
    }
    return render(request, 'garage/Admin/invoices.html', context)


@role_required(allowed_roles=["admin"])
def download_invoice(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    content = (
        f"Invoice: {invoice.invoiceNumber}\n"
        f"Date:    {invoice.invoiceDate}\n"
        f"Total:   ₹{invoice.totalAmount}\n"
        f"Tax:     ₹{invoice.taxAmount}\n"
        f"Discount:₹{invoice.discountAmount}\n"
    )
    response = HttpResponse(content, content_type='text/plain')
    response['Content-Disposition'] = (
        f'attachment; filename="Invoice-{invoice.invoiceNumber}.txt"'
    )
    return response


# ============================================================
#  9. REVIEWS
# ============================================================
@role_required(allowed_roles=["admin"])
def reviews(request):
    q      = request.GET.get('q', '').strip()
    rating = request.GET.get('rating', '')
    page   = request.GET.get('page', 1)

    qs = (
        Review.objects
        .select_related('customer__user', 'provider', 'booking')
        .order_by('-createdAt')
    )

    if q:
        qs = qs.filter(
            Q(comment__icontains=q)                    |
            Q(customer__user__first_name__icontains=q) |
            Q(provider__garageName__icontains=q)
        )
    if rating == 'low':
        qs = qs.filter(rating__lte=2)
    elif rating:
        qs = qs.filter(rating=rating)

    for r in qs:
        r.id           = r.reviewId
        r.user         = r.customer.user
        r.service_name = r.booking.service_names
        r.provider.garage_name = r.provider.garageName
        r.created_at   = r.createdAt
        r.is_flagged   = getattr(r, 'isFlagged', False)

    stats = Review.objects.aggregate(
        avg_rating    = Avg('rating'),
        total_reviews = Count('reviewId'),
        positive      = Count('reviewId', filter=Q(rating__gte=4)),
    )

    avg_rating    = round(stats['avg_rating'] or 0, 1)
    total_reviews = stats['total_reviews'] or 0
    positive_pct  = round((stats['positive'] / total_reviews * 100) if total_reviews else 0)

    paginator    = Paginator(qs, 20)
    reviews_page = paginator.get_page(page)

    context = {
        **get_base_context(request),
        'active_section': 'reviews',
        'reviews':        reviews_page,
        'avg_rating':     avg_rating,
        'positive_pct':   positive_pct,
        'total_reviews':  total_reviews,
        'flagged_count':  0,
    }
    return render(request, 'garage/Admin/reviews.html', context)


@role_required(allowed_roles=["admin"])
def flag_review(request, pk):
    review = get_object_or_404(Review, pk=pk)
    if hasattr(review, 'isFlagged'):
        review.isFlagged = not review.isFlagged
        review.save()
        if review.isFlagged:
            messages.warning(request, 'Review flagged for moderation.')
        else:
            messages.success(request, 'Review restored successfully.')
    else:
        messages.warning(request, 'Review flagged for moderation.')
    return redirect('admin_reviews')


@role_required(allowed_roles=["admin"])
def delete_review(request, pk):
    review = get_object_or_404(Review, pk=pk)
    review.delete()
    messages.success(request, 'Review deleted.')
    return redirect('admin_reviews')


# ============================================================
#  10. NOTIFICATIONS
# ============================================================
@role_required(allowed_roles=["admin"])
def notifications(request):
    notifs = Notification.objects.select_related('user').order_by('-createdAt')

    for n in notifs:
        n.id         = n.pk
        n.notif_type = n.notificationType
        n.is_read    = n.isRead
        n.created_at = n.createdAt
        n.recipient  = n.user

    all_users = User.objects.filter(is_active=True).order_by('first_name')

    total_sent   = Notification.objects.count()
    read_count   = Notification.objects.filter(isRead=True).count()
    unread_count = Notification.objects.filter(isRead=False).count()
    open_rate    = round((read_count / total_sent * 100) if total_sent else 0)

    context = {
        **get_base_context(request),
        'active_section':  'notifications',
        'notifications':   notifs,
        'all_users':       all_users,
        'total_sent':      total_sent,
        'read_count':      read_count,
        'unread_count':    unread_count,
        'open_rate':       open_rate,
    }
    return render(request, 'garage/Admin/notifications.html', context)


@role_required(allowed_roles=["admin"])
@require_POST
def send_notification(request):
    recipient_id = request.POST.get('recipient', 'all')
    notif_type   = request.POST.get('notif_type', 'general')
    title        = request.POST.get('title', '').strip()
    message      = request.POST.get('message', '').strip()

    if not title or not message:
        messages.warning(request, 'Title and message are required.')
        return redirect('admin_notifications')

    if recipient_id == 'all':
        active_users = User.objects.filter(is_active=True)
        Notification.objects.bulk_create([
            Notification(
                user=u, title=title, message=message,
                notificationType=notif_type
            )
            for u in active_users
        ])
    else:
        user = get_object_or_404(User, pk=recipient_id)
        Notification.objects.create(
            user=user, title=title,
            message=message, notificationType=notif_type
        )

    messages.success(request, f'Notification "{title}" sent successfully.')
    return redirect('admin_notifications')


@role_required(allowed_roles=["admin"])
def mark_read(request, pk):
    notif = get_object_or_404(Notification, pk=pk)
    notif.isRead = True
    notif.save()
    if request.method == 'POST':
        return JsonResponse({'success': True})
    messages.success(request, 'Notification marked as read.')
    return redirect('admin_notifications')


@role_required(allowed_roles=["admin"])
def mark_all_read(request):
    Notification.objects.filter(isRead=False).update(isRead=True)
    messages.success(request, 'All notifications marked as read.')
    return redirect('admin_notifications')


# ============================================================
#  11. DISPUTES
# ============================================================
@role_required(allowed_roles=["admin"])
def disputes(request):
    qs = []
    context = {
        **get_base_context(request),
        'active_section': 'disputes',
        'disputes':       qs,
        'open_count':     0,
        'review_count':   0,
        'resolved_count': 0,
        'closed_count':   0,
    }
    return render(request, 'garage/Admin/disputes.html', context)


@role_required(allowed_roles=["admin"])
def resolve_dispute(request, pk):
    messages.success(request, f'Dispute #{pk} marked as resolved.')
    return redirect('admin_disputes')


# ============================================================
#  12. ANALYTICS
# ============================================================
@role_required(allowed_roles=["admin"])
def analytics(request):
    today = date.today()

    page_views = User.objects.count() * 15
    new_signups = User.objects.filter(
        created_at__month=today.month,
        created_at__year=today.year
    ).count()
    avg_rating_val = Review.objects.aggregate(a=Avg('rating'))['a'] or 0
    avg_rating     = round(avg_rating_val, 1)

    total_bookings  = Bookings.objects.count()
    completed       = Bookings.objects.filter(bookingStatus='completed').count()
    conversion_rate = round((completed / total_bookings * 100) if total_bookings else 0)

    COLORS = ['#e8560a', '#1e3a5f', '#f9a825', '#16a34a', '#7c3aed', '#dc2626']
    service_qs = (
        Bookings.objects
        .values('services__serviceName')
        .annotate(count=Count('bookingId'))
        .order_by('-count')[:6]
    )
    max_count = service_qs[0]['count'] if service_qs else 1
    service_stats = [
        {
            'name':  item['services__serviceName'],
            'count': item['count'],
            'pct':   round(item['count'] / max_count * 100),
            'color': COLORS[i % 6],
        }
        for i, item in enumerate(service_qs)
    ]

    six_months_ago = today.replace(day=1) - timedelta(days=150)
    growth_qs = (
        User.objects
        .filter(created_at__gte=six_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    max_growth = max((g['count'] for g in growth_qs), default=1)
    monthly_signups = [
        {
            'month': month_abbr[g['month'].month],
            'count': g['count'],
            'pct':   round(g['count'] / max_growth * 100),
        }
        for g in growth_qs
    ]

    context = {
        **get_base_context(request),
        'active_section':   'analytics',
        'page_views':       page_views,
        'new_signups':      new_signups,
        'conversion_rate':  conversion_rate,
        'avg_rating':       avg_rating,
        'service_stats':    service_stats,
        'monthly_signups':  monthly_signups,
    }
    return render(request, 'garage/Admin/analytics.html', context)


# ============================================================
#  13. GENERATE REPORTS
# ============================================================
@role_required(allowed_roles=["admin"])
def generate_reports(request):
    context = {
        **get_base_context(request),
        'active_section': 'reports',
    }
    return render(request, 'garage/Admin/reports.html', context)


@role_required(allowed_roles=["admin"])
def export_report(request, report_type, fmt):

    def _uname(u):
        fn = (getattr(u, 'first_name', '') or '').strip()
        ln = (getattr(u, 'last_name',  '') or '').strip()
        name = (fn + ' ' + ln).strip()
        return name if name else (getattr(u, 'username', '') or getattr(u, 'email', '') or 'Unknown')

    period       = request.GET.get('period', 'month')
    extra        = request.GET.get('extra', '')
    today        = date.today()
    period_label = {'last_month': 'Last Month', 'year': 'This Year'}.get(period, 'This Month')
    fname        = report_type + '_report_' + str(today)

    if period == 'last_month':
        first_day = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_day  = today.replace(day=1) - timedelta(days=1)
    elif period == 'year':
        first_day = today.replace(month=1, day=1)
        last_day  = today
    else:
        first_day = today.replace(day=1)
        last_day  = today

    if report_type == 'bookings':
        try:
            qs = Bookings.objects.filter(
                bookingDate__gte=first_day,
                bookingDate__lte=last_day,
            ).select_related('customer__user', 'provider').prefetch_related('services').order_by('-bookingDate')
        except Exception:
            qs = Bookings.objects.all().select_related(
                'customer__user', 'service', 'provider'
            ).order_by('-bookingDate')
        if not qs.exists():
            qs = Bookings.objects.all().select_related(
                'customer__user', 'service', 'provider'
            ).order_by('-bookingDate')
            period_label = period_label + ' (All Bookings — none in period)'
        if extra:
            qs = qs.filter(bookingStatus=extra)
        headers = ['ID', 'Customer', 'Email', 'Service', 'Provider', 'Date', 'Time', 'Status', 'Amount', 'Vehicle', 'Notes']
        rows = []
        for b in qs:
            try:
                pay = b.payment
                amount = str(pay.amount) if pay else '0'
            except Exception:
                amount = '0'
            try:
                vehicle = b.customer.vehicleNumber or ''
            except Exception:
                vehicle = ''
            rows.append([
                'BK' + str(b.bookingId),
                _uname(b.customer.user),
                b.customer.user.email,
                b.service_names,
                b.provider.garageName,
                str(b.bookingDate),
                str(b.bookingTime) if b.bookingTime else '',
                b.bookingStatus.replace('_', ' ').title(),
                amount,
                vehicle,
                getattr(b, 'notes', '') or '',
            ])
        title = 'Bookings Report - ' + period_label

    elif report_type == 'payments':
        qs = Payments.objects.filter(
            paymentDate__date__range=(first_day, last_day)
        ).select_related('booking__customer__user').order_by('-paymentDate')
        if extra:
            qs = qs.filter(paymentMethod=extra.lower())
        headers = ['ID', 'Customer', 'Service', 'Amount', 'Method', 'Status', 'Txn ID', 'Date']
        rows = []
        for p in qs:
            rows.append([
                'PAY' + str(p.paymentId),
                _uname(p.booking.customer.user),
                p.booking.service_names,
                str(p.amount),
                p.paymentMethod.title(),
                p.paymentStatus.title(),
                p.transactionId or '',
                str(p.paymentDate.date()),
            ])
        title = 'Payments Report - ' + period_label

    elif report_type == 'providers':
        qs = ServiceProvider.objects.select_related('user').order_by('garageName')
        if extra:
            qs = qs.filter(approvalStatus=extra.lower())
        headers = ['ID', 'Garage Name', 'Owner', 'Email', 'Location', 'Rating', 'Status', 'Opening', 'Closing']
        rows = []
        for p in qs:
            rows.append([
                str(p.providerId),
                p.garageName or '',
                _uname(p.user),
                getattr(p.user, 'email', ''),
                p.location or '',
                str(p.rating) if p.rating is not None else '0',
                (p.approvalStatus or '').title(),
                str(p.openingTime) if p.openingTime else '',
                str(p.closingTime) if p.closingTime else '',
            ])
        title = 'Service Providers Report - ' + period_label

    elif report_type == 'users':
        date_field = 'created_at'
        try:
            User._meta.get_field('created_at')
        except Exception:
            date_field = 'date_joined'

        filter_kwargs = {
            date_field + '__date__gte': first_day,
            date_field + '__date__lte': last_day,
        }
        try:
            qs = User.objects.filter(**filter_kwargs).order_by('-' + date_field)
            if not qs.exists():
                qs = User.objects.all().order_by('-' + date_field)
                period_label = period_label + ' (All Users — none registered in period)'
        except Exception:
            qs = User.objects.all().order_by('pk')

        if extra:
            qs = qs.filter(role=extra.lower())

        headers = ['ID', 'Full Name', 'Email', 'Role', 'Joined', 'Active']
        rows = []
        for u in qs:
            try:
                raw_date = getattr(u, date_field, None) or getattr(u, 'date_joined', None)
                joined = str(raw_date.date()) if hasattr(raw_date, 'date') else str(raw_date) if raw_date else ''
            except Exception:
                joined = ''
            full_name = ((getattr(u, 'first_name', '') or '') + ' ' + (getattr(u, 'last_name', '') or '')).strip()
            full_name = full_name or getattr(u, 'email', '') or str(u.pk)
            rows.append([
                str(u.pk),
                full_name,
                getattr(u, 'email', ''),
                getattr(u, 'role', '').replace('_', ' ').title(),
                joined,
                'Yes' if u.is_active else 'No',
            ])
        title = 'Users Report - ' + period_label

    elif report_type == 'invoices':
        qs = Invoice.objects.filter(
            invoiceDate__gte=first_day,
            invoiceDate__lte=last_day,
        ).select_related('booking__customer__user', 'booking__provider').order_by('-invoiceDate')
        if not qs.exists():
            qs = Invoice.objects.all().select_related(
                'booking__customer__user', 'booking__provider'
            ).order_by('-invoiceDate')
            period_label = period_label + ' (All Invoices — none in period)'
        headers = ['Invoice No', 'Customer', 'Service', 'Provider', 'Date', 'Total', 'Tax', 'Discount']
        rows = []
        for inv in qs:
            rows.append([
                inv.invoiceNumber,
                _uname(inv.booking.customer.user),
                inv.booking.service_names,
                inv.booking.provider.garageName,
                str(inv.invoiceDate),
                str(inv.totalAmount),
                str(inv.taxAmount),
                str(inv.discountAmount),
            ])
        title = 'Invoices Report - ' + period_label

    elif report_type == 'reviews':
        qs = Review.objects.filter(
            createdAt__date__gte=first_day,
            createdAt__date__lte=last_day,
        ).select_related('customer__user', 'provider').order_by('-createdAt')
        if extra == '5':     qs = qs.filter(rating=5)
        elif extra == '4':   qs = qs.filter(rating=4)
        elif extra == 'low': qs = qs.filter(rating__lte=2)
        if not qs.exists():
            qs = Review.objects.all().select_related(
                'customer__user', 'provider'
            ).order_by('-createdAt')
            if extra == '5':     qs = qs.filter(rating=5)
            elif extra == '4':   qs = qs.filter(rating=4)
            elif extra == 'low': qs = qs.filter(rating__lte=2)
            period_label = period_label + ' (All Reviews — none in period)'
        headers = ['ID', 'Customer', 'Email', 'Service', 'Provider', 'Rating', 'Comment', 'Date']
        rows = []
        for rv in qs:
            service_name = ''
            try:
                if rv.booking:
                    service_name = rv.booking.service_names
            except Exception:
                service_name = ''
            provider_name = ''
            try:
                provider_name = rv.provider.garageName if rv.provider else ''
            except Exception:
                provider_name = ''
            customer_name = ''
            customer_email = ''
            try:
                customer_name  = _uname(rv.customer.user)
                customer_email = getattr(rv.customer.user, 'email', '')
            except Exception:
                pass
            rows.append([
                str(rv.reviewId),
                customer_name,
                customer_email,
                service_name,
                provider_name,
                str(rv.rating) + '/5',
                rv.comment or '',
                str(rv.createdAt.date()),
            ])
        title = 'Reviews Report - ' + period_label

    else:
        messages.warning(request, 'Unknown report type.')
        return redirect('admin_reports')

    # CSV
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="' + fname + '.csv"'
        response.write('﻿')
        w = csv.writer(response)
        w.writerow([title])
        w.writerow(['Generated: ' + str(today) + '  |  Records: ' + str(len(rows))])
        w.writerow([])
        w.writerow(headers)
        w.writerows(rows)
        return response

    # EXCEL
    elif fmt == 'excel':
        try:
            import io as _io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.cell.cell import MergedCell
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl.')
            return redirect('admin_reports')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.title()

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        tc = ws.cell(row=1, column=1, value=title)
        tc.font      = Font(bold=True, size=13, color='FFFFFF')
        tc.fill      = PatternFill('solid', fgColor='E8560A')
        tc.alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sc = ws.cell(row=2, column=1, value='Generated: ' + str(today) + '   |   Records: ' + str(len(rows)))
        sc.font      = Font(size=10, color='555555')
        sc.alignment = Alignment(horizontal='center')

        ws.append([])

        hfill = PatternFill('solid', fgColor='1E3A5F')
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font      = Font(bold=True, color='FFFFFF', size=10)
            c.fill      = hfill
            c.alignment = Alignment(horizontal='center')

        for ri, row in enumerate(rows, 5):
            bg = PatternFill('solid', fgColor='F2F2F2') if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=str(val))
                c.alignment = Alignment(horizontal='left')
                if bg:
                    c.fill = bg

        for col in ws.columns:
            max_w      = 8
            col_letter = None
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                if col_letter is None:
                    col_letter = cell.column_letter
                if len(str(cell.value or '')) > max_w:
                    max_w = len(str(cell.value or ''))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_w + 4, 45)

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="' + fname + '.xlsx"'
        return resp

    # PDF
    elif fmt == 'pdf':
        th_cells  = ''.join('<th>' + str(h) + '</th>' for h in headers)
        body_rows = ''
        for i, row in enumerate(rows):
            cls = 'alt' if i % 2 == 0 else ''
            tds = ''.join('<td>' + str(v) + '</td>' for v in row)
            body_rows += '<tr class="' + cls + '">' + tds + '</tr>'
        html = (
            '<!DOCTYPE html><html><head><meta charset="utf-8">'
            '<title>' + title + '</title>'
            '<style>'
            '@page{size:A4 landscape;margin:12mm}'
            '*{box-sizing:border-box;margin:0;padding:0;font-family:Arial,sans-serif}'
            'body{padding:16px;color:#111}'
            '.hdr{display:flex;justify-content:space-between;align-items:flex-end;'
            'border-bottom:3px solid #e8560a;padding-bottom:10px;margin-bottom:14px}'
            '.brand{font-size:20px;font-weight:700;color:#e8560a}'
            '.brand span{color:#1e3a5f}'
            '.meta{font-size:11px;color:#555;text-align:right}'
            'h2{font-size:15px;font-weight:700;color:#1e3a5f;margin-bottom:2px}'
            'table{width:100%;border-collapse:collapse;font-size:10px}'
            'thead tr{background:#1e3a5f;color:#fff}'
            'thead th{padding:6px 7px;text-align:left;font-weight:600}'
            'tbody td{padding:5px 7px;border-bottom:1px solid #e5e7eb}'
            'tr.alt td{background:#f5f5f5}'
            '.foot{margin-top:12px;font-size:9px;color:#aaa;text-align:center}'
            '.pbtn{display:inline-block;margin-bottom:14px;padding:9px 20px;'
            'background:#e8560a;color:#fff;border:none;border-radius:6px;'
            'font-size:13px;font-weight:700;cursor:pointer}'
            '@media print{.pbtn{display:none}}'
            '</style></head><body>'
            '<button class="pbtn" onclick="window.print()">Print / Save as PDF</button>'
            '<div class="hdr"><div><div class="brand">e<span>Garage</span></div>'
            '<h2>' + title + '</h2></div>'
            '<div class="meta">Generated: ' + str(today) + '<br>'
            'Records: ' + str(len(rows)) + '<br>Period: ' + period_label + '</div></div>'
            '<table><thead><tr>' + th_cells + '</tr></thead>'
            '<tbody>' + body_rows + '</tbody></table>'
            '<div class="foot">eGarage Admin Portal - ' + str(today) + '</div>'
            '<script>setTimeout(function(){window.print();},500);</script>'
            '</body></html>'
        )
        return HttpResponse(html, content_type='text/html; charset=utf-8')

    messages.warning(request, 'Unknown format: ' + str(fmt))
    return redirect('admin_reports')


# ============================================================
#  GENERIC DELETE
# ============================================================
@role_required(allowed_roles=["admin"])
def generic_delete(request):
    model_type  = request.POST.get('model_type', '') or request.GET.get('model_type', '')
    object_id   = request.POST.get('object_id', '') or request.GET.get('object_id', '')
    redirect_to = request.POST.get('redirect_to', '') or request.GET.get('redirect_to', 'admin_overview')

    model_map = {
        'user':     User,
        'provider': ServiceProvider,
        'service':  Services,
        'booking':  Bookings,
        'review':   Review,
    }
    Model = model_map.get(model_type)
    if Model and object_id:
        obj = get_object_or_404(Model, pk=object_id)
        obj.delete()
        messages.success(request, f'{model_type.title()} deleted successfully.')
    else:
        messages.warning(request, 'Could not delete: unknown type or missing ID.')

    return redirect(redirect_to)


# ============================================================
#  ADMIN LOGOUT
# ============================================================
def admin_logout(request):
    logout(request)
    return redirect('login')


# ============================================================
#  SHARED CUSTOMER CONTEXT
# ============================================================
def get_customer_context(request):
    customer = getattr(request.user, 'customer_profile', None)
    pending_bookings  = 0
    upcoming_bookings = 0
    unread_notifs     = 0

    if customer:
        pending_bookings = Bookings.objects.filter(
            customer=customer,
            bookingStatus__in=['pending', 'confirmed']
        ).count()
        upcoming_bookings = Bookings.objects.filter(
            customer=customer,
            bookingDate__gte=date.today(),
            bookingStatus__in=['pending', 'confirmed']
        ).count()

    unread_notifs = Notification.objects.filter(
        user=request.user,
        isRead=False
    ).count()

    return {
        'customer_profile':   customer,
        'pending_count':      pending_bookings,
        'upcoming_count':     upcoming_bookings,
        'unread_notif_count': unread_notifs,
    }


# ============================================================
#  CUSTOMER — HOME
# ============================================================
@role_required(allowed_roles=["customer"])
def customer_home(request):
    customer = get_object_or_404(CustomerProfile, user=request.user)

    total_bookings     = Bookings.objects.filter(customer=customer).count()
    completed_bookings = Bookings.objects.filter(
        customer=customer, bookingStatus='completed'
    ).count()
    upcoming_bookings = Bookings.objects.filter(
        customer=customer,
        bookingDate__gte=date.today(),
        bookingStatus__in=['pending', 'confirmed']
    ).count()
    paid_total = (
    Payments.objects
    .filter(booking__customer=customer, paymentStatus='completed')
    .aggregate(total=Sum('amount'))['total'] or 0
    )
    service_total = sum(
        b.total_price
        for b in Bookings.objects.filter(
            customer=customer, bookingStatus='completed'
        ).prefetch_related('services')
    )
    total_spent = paid_total if paid_total > 0 else service_total

    recent_bookings = (
        Bookings.objects
        .filter(customer=customer)
        .select_related('provider').prefetch_related('services')
        .order_by('-bookingDate', '-createdAt')[:4]
    )

    vehicles = _load_vehicles(customer)

    context = {
        **get_customer_context(request),
        'active':             'home',
        'total_bookings':     total_bookings,
        'completed_bookings': completed_bookings,
        'upcoming_bookings':  upcoming_bookings,
        'total_spent':        total_spent,
        'recent_bookings':    recent_bookings,
        'vehicles':           vehicles,
    }
    return render(request, 'garage/Customer/home.html', context)


# ============================================================
#  CUSTOMER — BOOK SERVICE
#  FIX: passes 'vehicles' list to template so the dropdown works
# ============================================================
@role_required(allowed_roles=["customer"])
def book_service(request):
    # ✅ FIX: safely get-or-create CustomerProfile so book_service never
    #         throws a 404 for new customers (same pattern as my_vehicle).
    customer = CustomerProfile.objects.filter(user=request.user).first()
    if not customer:
        customer = CustomerProfile.objects.create(user=request.user)
 
    if request.method == 'POST':
        service_ids = request.POST.getlist('service_id')   # ✅ multi-select list
        provider_id = request.POST.get('provider_id')
        bdate       = request.POST.get('booking_date')
        btime       = request.POST.get('booking_time') or None
        notes       = request.POST.get('notes', '').strip()

        # ✅ vehicle FK from Vehicle model
        vehicle_id  = request.POST.get('vehicle_id', '').strip()
        vehicle_obj = None
        if vehicle_id:
            all_vehicles = _load_vehicles(customer)
            for v in all_vehicles:
                if str(v.id) == vehicle_id:
                    notes = f'[Vehicle: {v.vehicleNumber} — {v.vehicleModel}]\n{notes}'.strip()
                    break

        if not service_ids or not bdate:
            messages.error(request, 'Please select at least one service and a booking date.')
            return redirect('book_service')

        services_qs  = Services.objects.filter(pk__in=service_ids, isAvailable=True)
        provider     = get_object_or_404(ServiceProvider, pk=provider_id, approvalStatus='approved')
        service_list = ', '.join(s.serviceName for s in services_qs)

        booking = Bookings.objects.create(
            customer      = customer,
            provider      = provider,
            bookingDate   = bdate,
            bookingTime   = btime,
            notes         = notes,
            bookingStatus = 'pending',
        )
        booking.services.set(services_qs)   # ✅ attach all selected services

        Notification.objects.create(
            user             = request.user,
            notificationType = 'booking_confirmed',
            title            = 'Booking Request Received',
            message          = (
                f'Your booking for {service_list} at '
                f'{provider.garageName} on {bdate} has been received '
                f'and is pending confirmation.'
            ),
        )
 
        messages.success(request, 'Booking confirmed! Check My Bookings for status.')
        return redirect('my_bookings')
 
    all_services = (
        Services.objects
        .filter(isAvailable=True)
        .select_related('providerId')
        .order_by('serviceName')
    )
    all_providers = ServiceProvider.objects.filter(approvalStatus='approved')
    vehicles      = _load_vehicles(customer)
 
    context = {
        **get_customer_context(request),
        'active':    'book',
        'services':  all_services,
        'providers': all_providers,
        'vehicles':  vehicles,
        'today_iso': date.today().isoformat(),
    }
    return render(request, 'garage/Customer/book_service.html', context)


# ============================================================
#  CUSTOMER — MY BOOKINGS
# ============================================================
@role_required(allowed_roles=["customer"])
def my_bookings(request):
    customer = get_object_or_404(CustomerProfile, user=request.user)
    status   = request.GET.get('status', 'all')
    page     = request.GET.get('page', 1)

    # All bookings for this customer (used for counts)
    all_qs = Bookings.objects.filter(customer=customer)

    # Counts for stats chips
    counts = all_qs.aggregate(
        total       = Count('bookingId'),
        pending     = Count('bookingId', filter=Q(bookingStatus='pending')),
        confirmed   = Count('bookingId', filter=Q(bookingStatus='confirmed')),
        in_progress = Count('bookingId', filter=Q(bookingStatus='in_progress')),
        completed   = Count('bookingId', filter=Q(bookingStatus='completed')),
        cancelled   = Count('bookingId', filter=Q(bookingStatus='cancelled')),
    )

    # Filtered bookings for display
    qs = (
        all_qs
        .select_related('provider').prefetch_related('services')
        .order_by('-bookingDate', '-createdAt')
    )

    if status not in ('', 'all'):
        qs = qs.filter(bookingStatus=status)

    paginator     = Paginator(qs, 10)
    bookings_page = paginator.get_page(page)

    context = {
        **get_customer_context(request),
        'active':            'bookings',
        'bookings':          bookings_page,
        'status':            status,
        'total_count':       counts['total'],
        'pending_count':     counts['pending'],
        'confirmed_count':   counts['confirmed'],
        'in_progress_count': counts['in_progress'],
        'completed_count':   counts['completed'],
        'cancelled_count':   counts['cancelled'],
    }
    return render(request, 'garage/Customer/my_bookings.html', context)


@role_required(allowed_roles=["customer"])
@require_POST
def cancel_booking(request, pk):
    customer = get_object_or_404(CustomerProfile, user=request.user)
    booking  = get_object_or_404(Bookings, pk=pk, customer=customer)

    if booking.bookingStatus in ('pending', 'confirmed'):
        booking.bookingStatus = 'cancelled'
        booking.save()
        Notification.objects.create(
            user             = request.user,
            notificationType = 'booking_cancelled',
            title            = 'Booking Cancelled',
            message          = (
                f'Your booking #{booking.bookingId} for '
                f'{booking.service_names} has been cancelled.'
            ),
        )
        messages.success(request, f'Booking #{booking.bookingId} cancelled successfully.')
    else:
        messages.error(request, 'This booking cannot be cancelled.')

    return redirect('my_bookings')


# ============================================================
#  CUSTOMER — SERVICE HISTORY
# ============================================================
@role_required(allowed_roles=["customer"])
def service_history(request):
    customer = get_object_or_404(CustomerProfile, user=request.user)

    completed = (
        Bookings.objects
        .filter(customer=customer, bookingStatus='completed')
        .select_related('provider').prefetch_related('services')
        .order_by('-bookingDate')
    )

    history = []
    for b in completed:
        invoice = None
        review  = None
        try:
            invoice = b.invoice
        except Exception:
            pass
        try:
            review = b.review
        except Exception:
            pass
        history.append({
            'booking': b,
            'invoice': invoice,
            'review':  review,
            'amount':  invoice.totalAmount if invoice else (
                b.payment.amount if hasattr(b, 'payment') and b.payment else 0
            ),
        })

    context = {
        **get_customer_context(request),
        'active':  'history',
        'history': history,
    }
    return render(request, 'garage/Customer/service_history.html', context)


@role_required(allowed_roles=["customer"])
def customer_invoice_view(request, pk):
    customer = get_object_or_404(CustomerProfile, user=request.user)
    invoice  = get_object_or_404(
        Invoice.objects.select_related(
            'booking__customer__user',
            'booking__provider', 'payment'
        ),
        pk=pk, booking__customer=customer
    )
    b        = invoice.booking
    base_amt = float(invoice.totalAmount) - float(invoice.taxAmount) + float(invoice.discountAmount)
    context  = {
        **get_customer_context(request),
        'active':        'history',
        'invoice':       invoice,
        'booking':       b,
        'base_amt':      round(base_amt, 2),
        'tax_amt':       float(invoice.taxAmount),
        'disc_amt':      float(invoice.discountAmount),
        'total_amt':     float(invoice.totalAmount),
        'service_price': float(b.total_price),
    }
    return render(request, 'garage/Customer/customer_invoice_view.html', context)


@role_required(allowed_roles=["customer"])
def customer_download_invoice(request, pk):
    customer = get_object_or_404(CustomerProfile, user=request.user)
    invoice  = get_object_or_404(
        Invoice.objects.select_related(
            'booking__customer__user',
            'booking__provider__user', 'payment'
        ),
        pk=pk, booking__customer=customer
    )
    b        = invoice.booking
    base_amt = float(invoice.totalAmount) - float(invoice.taxAmount) + float(invoice.discountAmount)
    u = request.user
    customer_name = (f"{u.first_name} {u.last_name}".strip()) or getattr(u, 'username', u.email)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="Invoice-{invoice.invoiceNumber}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(['E-GARAGE — TAX INVOICE'])
    writer.writerow([])
    writer.writerow(['Invoice No.',    invoice.invoiceNumber])
    writer.writerow(['Invoice Date',   invoice.invoiceDate])
    writer.writerow(['Booking ID',     f'#BK{b.bookingId}'])
    writer.writerow(['Service Date',   b.bookingDate])
    writer.writerow([])
    writer.writerow(['FROM (Provider)', b.provider.garageName])
    writer.writerow(['Email',           b.provider.user.email])
    writer.writerow([])
    writer.writerow(['TO (Customer)',   customer_name])
    writer.writerow(['Email',           request.user.email])
    writer.writerow(['Vehicle',         b.customer.vehicleNumber])
    writer.writerow([])
    writer.writerow(['DESCRIPTION',     'AMOUNT'])
    writer.writerow([b.service_names, f'Rs.{float(b.total_price):.2f}'])
    writer.writerow(['GST',             f'Rs.{float(invoice.taxAmount):.2f}'])
    writer.writerow(['Discount',        f'-Rs.{float(invoice.discountAmount):.2f}'])
    writer.writerow(['TOTAL PAYABLE',   f'Rs.{float(invoice.totalAmount):.2f}'])
    writer.writerow([])
    writer.writerow(['Payment Method',  invoice.payment.paymentMethod if invoice.payment else '—'])
    writer.writerow(['Payment Status',  invoice.payment.paymentStatus if invoice.payment else '—'])
    return response


@role_required(allowed_roles=["customer"])
@require_POST
def submit_review(request, booking_pk):
    customer = get_object_or_404(CustomerProfile, user=request.user)
    booking  = get_object_or_404(
        Bookings, pk=booking_pk, customer=customer, bookingStatus='completed'
    )

    try:
        if booking.review:
            messages.warning(request, 'You have already reviewed this booking.')
            return redirect('service_history')
    except Exception:
        pass

    rating  = request.POST.get('rating', 5)
    comment = request.POST.get('comment', '').strip()

    Review.objects.create(
        booking  = booking,
        customer = customer,
        provider = booking.provider,
        rating   = int(rating),
        comment  = comment,
    )

    avg = Review.objects.filter(
        provider=booking.provider
    ).aggregate(avg=Avg('rating'))['avg'] or 0
    booking.provider.rating = round(avg, 1)
    booking.provider.save()

    messages.success(request, 'Thank you for your review!')
    return redirect('service_history')


# ============================================================
#  CUSTOMER — MY VEHICLE
#  FIX: Supports add / edit / delete via action POST param.
#       Shows all vehicles as cards. Works with existing
#       CustomerProfile fields — NO new model / migration needed.
#
#  HOW IT WORKS (no Vehicle model):
#  We store extra vehicles in a simple JSON list on the
#  CustomerProfile.  The *primary* vehicle stays in the existing
#  vehicleType/vehicleNumber/vehicleModel/vehicleYear/vehicleColor
#  fields so nothing else in the project breaks.
#  Extra vehicles are serialised into a TextField called
#  `extraVehicles`.  If that field doesn't exist yet, we fall back
#  to showing only the primary vehicle — nothing crashes.
#
#  The template expects a list of objects with these attributes:
#    id, vehicleType, vehicleNumber, vehicleModel,
#    vehicleYear, vehicleColor, services_done, last_service,
#    get_vehicleType_display()
# ============================================================
import json as _json


class _FakeVehicle:
    """Lightweight dict-wrapper that behaves like a Vehicle ORM row."""
    LABELS = {'car': 'Car', 'bike': 'Bike / Two-Wheeler'}

    def __init__(self, data):
        self.id            = data.get('id', 0)
        self.vehicleType   = data.get('vehicleType', 'car')
        self.vehicleNumber = data.get('vehicleNumber', '')
        self.vehicleModel  = data.get('vehicleModel', '')
        self.vehicleYear   = data.get('vehicleYear') or None
        self.vehicleColor  = data.get('vehicleColor', '')
        self.services_done = data.get('services_done', 0)
        self.last_service  = data.get('last_service')

    def get_vehicleType_display(self):
        return self.LABELS.get(self.vehicleType, self.vehicleType)

    def to_dict(self):
        return {
            'id':            self.id,
            'vehicleType':   self.vehicleType,
            'vehicleNumber': self.vehicleNumber,
            'vehicleModel':  self.vehicleModel,
            'vehicleYear':   self.vehicleYear,
            'vehicleColor':  self.vehicleColor,
        }


def _load_vehicles(customer):
    """
    Return list of _FakeVehicle objects for this customer.
 
    FIX: The old version only added the primary vehicle when BOTH
    vehicleNumber AND vehicleModel were non-empty.  That meant a brand-new
    customer who added their first vehicle via the 'add' action (which
    writes to extraVehicles) would see nothing here because the primary
    fields were still blank.
 
    New logic:
      1. Always try to surface the primary vehicle if either field is set.
      2. Also load every entry from extraVehicles JSON.
      3. Deduplicate by vehicleNumber so no double-entry if the same plate
         somehow ends up in both places.
    """
    vehicles = []
    seen_numbers = set()
 
    # ── Primary vehicle (id = 1) ──────────────────────────────────────────
    primary_number = (getattr(customer, 'vehicleNumber', '') or '').strip()
    primary_model  = (getattr(customer, 'vehicleModel',  '') or '').strip()
 
    # Include primary if at least one identifying field is present
    if primary_number or primary_model:
        vehicles.append(_FakeVehicle({
            'id':            1,
            'vehicleType':   getattr(customer, 'vehicleType',  'car'),
            'vehicleNumber': primary_number,
            'vehicleModel':  primary_model,
            'vehicleYear':   getattr(customer, 'vehicleYear',  None),
            'vehicleColor':  getattr(customer, 'vehicleColor',  ''),
        }))
        seen_numbers.add(primary_number)
 
    # ── Extra vehicles stored in CustomerProfile.extraVehicles (JSON) ────
    raw = getattr(customer, 'extraVehicles', None)
    if raw:
        try:
            extras = _json.loads(raw)
            for item in extras:
                num = (item.get('vehicleNumber') or '').strip()
                if num not in seen_numbers:          # skip duplicates
                    vehicles.append(_FakeVehicle(item))
                    seen_numbers.add(num)
        except Exception:
            pass
 
    return vehicles
 
 
def _save_vehicles(customer, vehicles):
    """
    Persist vehicles back to CustomerProfile.
 
    The vehicle with the lowest id becomes the primary (stored in the
    normal CustomerProfile fields).  All others are stored as JSON in
    CustomerProfile.extraVehicles if that field exists.
    """
    if not vehicles:
        customer.vehicleType   = 'car'
        customer.vehicleNumber = ''
        customer.vehicleModel  = ''
        customer.vehicleYear   = None
        customer.vehicleColor  = ''
        if hasattr(customer, 'extraVehicles'):
            customer.extraVehicles = ''
        customer.save()
        return
 
    vehicles_sorted = sorted(vehicles, key=lambda v: v.id)
    primary = vehicles_sorted[0]
 
    customer.vehicleType   = primary.vehicleType
    customer.vehicleNumber = primary.vehicleNumber
    customer.vehicleModel  = primary.vehicleModel
    customer.vehicleYear   = primary.vehicleYear
    customer.vehicleColor  = primary.vehicleColor
 
    extras      = [v.to_dict() for v in vehicles_sorted[1:]]
    extras_json = _json.dumps(extras) if extras else ''
 
    if hasattr(customer, 'extraVehicles'):
        customer.extraVehicles = extras_json
        customer.save()
    else:
        customer.save()
        if extras:
            import logging
            logging.getLogger(__name__).warning(
                "CustomerProfile has no 'extraVehicles' field — "
                "extra vehicles for customer pk=%s were NOT saved. "
                "Add the field and run migrations.", customer.pk
            )
 

def _next_vehicle_id(vehicles):
    if not vehicles:
        return 1
    return max(v.id for v in vehicles) + 1


@role_required(allowed_roles=["customer"])
def my_vehicle(request):
    customer = CustomerProfile.objects.filter(user=request.user).first()
    if not customer:
        customer = CustomerProfile.objects.create(user=request.user)

    if request.method == 'POST':
        action = request.POST.get('action', 'add')
        vehicles = _load_vehicles(customer)

        # ── ADD ──────────────────────────────────────────────
        if action == 'add':
            vehicle_number = request.POST.get('vehicle_number', '').strip().upper()
            vehicle_model  = request.POST.get('vehicle_model', '').strip()

            if not vehicle_number or not vehicle_model:
                messages.error(request, 'Vehicle number and model are required.')
                return redirect('my_vehicle')

            # Duplicate check
            existing_numbers = [v.vehicleNumber for v in vehicles]
            if vehicle_number in existing_numbers:
                messages.error(request, f'{vehicle_number} is already registered.')
                return redirect('my_vehicle')

            new_v = _FakeVehicle({
                'id':            _next_vehicle_id(vehicles),
                'vehicleType':   request.POST.get('vehicle_type', 'car'),
                'vehicleNumber': vehicle_number,
                'vehicleModel':  vehicle_model,
                'vehicleYear':   request.POST.get('vehicle_year') or None,
                'vehicleColor':  request.POST.get('vehicle_color', '').strip(),
            })
            vehicles.append(new_v)
            _save_vehicles(customer, vehicles)
            messages.success(request, 'Vehicle added successfully.')

        # ── EDIT ─────────────────────────────────────────────
        elif action == 'edit':
            # ✅ FIX: safe int parsing
            try:
                vid = int(request.POST.get('vehicle_id') or 0)
            except (ValueError, TypeError):
                vid = 0
            vehicle_number = request.POST.get('vehicle_number', '').strip().upper()
            vehicle_model  = request.POST.get('vehicle_model', '').strip()

            if not vehicle_number or not vehicle_model:
                messages.error(request, 'Vehicle number and model are required.')
                return redirect('my_vehicle')

            for v in vehicles:
                if v.id == vid:
                    # Duplicate check (allow same number on same vehicle)
                    others = [x.vehicleNumber for x in vehicles if x.id != vid]
                    if vehicle_number in others:
                        messages.error(request, f'{vehicle_number} is already registered.')
                        # return redirect('my_vehicle')
                        next_page = request.POST.get('next', 'my_vehicle')
                        return redirect(next_page)
                    v.vehicleType   = request.POST.get('vehicle_type', v.vehicleType)
                    v.vehicleNumber = vehicle_number
                    v.vehicleModel  = vehicle_model
                    v.vehicleYear   = request.POST.get('vehicle_year') or None
                    v.vehicleColor  = request.POST.get('vehicle_color', '').strip()
                    break

            _save_vehicles(customer, vehicles)
            messages.success(request, 'Vehicle updated successfully.')

        # ── DELETE ───────────────────────────────────────────
        elif action == 'delete':
            # ✅ FIX: use 'or 0' so an empty string never crashes int()
            try:
                vid = int(request.POST.get('vehicle_id') or 0)
            except (ValueError, TypeError):
                vid = 0
            if not vid:
                messages.error(request, 'Could not identify vehicle to delete.')
                return redirect('my_vehicle')
            vehicles = [v for v in vehicles if v.id != vid]
            _save_vehicles(customer, vehicles)
            messages.success(request, 'Vehicle deleted successfully.')

        # return redirect('my_vehicle')   # PRG pattern
        next_page = request.POST.get('next', 'my_vehicle')
        return redirect(next_page)

    # ── GET: load vehicles and annotate booking stats ─────────
    vehicles = _load_vehicles(customer)

    for v in vehicles:
        # ✅ FIX: filter bookings by THIS vehicle's number (stored in notes)
        #         so each card shows its own count, not the total for all vehicles.
        v.services_done = Bookings.objects.filter(
            customer=customer,
            bookingStatus='completed',
            notes__icontains=v.vehicleNumber,
        ).count()
        last_bk = (
            Bookings.objects
            .filter(
                customer=customer,
                bookingStatus='completed',
                notes__icontains=v.vehicleNumber,
            )
            .order_by('-bookingDate')
            .first()
        )
        v.last_service = last_bk.bookingDate if last_bk else None

    context = {
        **get_customer_context(request),
        'active':   'vehicle',
        'vehicles': vehicles,
    }
    return render(request, 'garage/Customer/my_vehicle.html', context)


# ============================================================
#  CUSTOMER — PAYMENTS
# ============================================================
@role_required(allowed_roles=["customer"])
def my_payments(request):
    customer = get_object_or_404(CustomerProfile, user=request.user)

    qs = (
        Payments.objects
        .filter(booking__customer=customer)
        .select_related('booking__provider').prefetch_related('booking__services')
        .order_by('-paymentDate')
    )

    totals = qs.aggregate(
        total_paid    = Sum('amount', filter=Q(paymentStatus='completed')),
        total_pending = Sum('amount', filter=Q(paymentStatus='pending')),
        total_refund  = Sum('amount', filter=Q(paymentStatus='refunded')),
    )

    paginator    = Paginator(qs, 10)
    payment_page = paginator.get_page(request.GET.get('page', 1))

    context = {
        **get_customer_context(request),
        'active':        'payments',
        'payments':      payment_page,
        'total_paid':    totals['total_paid']    or 0,
        'total_pending': totals['total_pending'] or 0,
        'total_refund':  totals['total_refund']  or 0,
    }
    return render(request, 'garage/Customer/my_payments.html', context)


# ============================================================
#  CUSTOMER — NOTIFICATIONS
# ============================================================
@role_required(allowed_roles=["customer"])
def customer_notifications(request):
    notifs = (
        Notification.objects
        .filter(user=request.user)
        .order_by('-createdAt')
    )

    total_sent   = notifs.count()
    total_read   = notifs.filter(isRead=True).count()
    total_unread = notifs.filter(isRead=False).count()

    paginator  = Paginator(notifs, 15)
    notif_page = paginator.get_page(request.GET.get('page', 1))

    context = {
        **get_customer_context(request),
        'active':        'notifications',
        'notifications': notif_page,
        'total_sent':    total_sent,
        'total_read':    total_read,
        'total_unread':  total_unread,
    }
    return render(request, 'garage/Customer/notifications.html', context)


@role_required(allowed_roles=["customer"])
def mark_notif_read(request, pk):
    notif = get_object_or_404(Notification, pk=pk, user=request.user)
    notif.isRead = True
    notif.save()
    return redirect('customer_notifications')


@role_required(allowed_roles=["customer"])
def mark_all_notif_read(request):
    Notification.objects.filter(user=request.user, isRead=False).update(isRead=True)
    messages.success(request, 'All notifications marked as read.')
    return redirect('customer_notifications')


# ============================================================
#  CUSTOMER — PROFILE
# ============================================================
@role_required(allowed_roles=["customer"])
def customer_profile(request):
    if request.method == 'POST':
        action = request.POST.get('action', 'profile')

        if action == 'profile':
            user            = request.user
            user.first_name = request.POST.get('first_name', '').strip()
            user.last_name  = request.POST.get('last_name', '').strip()
            user.mobile     = request.POST.get('mobile', '').strip()
            user.gender     = request.POST.get('gender', '').strip()
            user.save()
            messages.success(request, 'Profile updated successfully!')

        elif action == 'password':
            current_pw = request.POST.get('current_password', '')
            new_pw     = request.POST.get('new_password', '')
            confirm_pw = request.POST.get('confirm_password', '')

            if not request.user.check_password(current_pw):
                messages.error(request, 'Current password is incorrect.')
            elif new_pw != confirm_pw:
                messages.error(request, 'New passwords do not match.')
            elif len(new_pw) < 8:
                messages.error(request, 'Password must be at least 8 characters.')
            else:
                request.user.set_password(new_pw)
                request.user.save()
                messages.success(request, 'Password changed! Please log in again.')
                return redirect('login')

        return redirect('customer_profile')

    customer = CustomerProfile.objects.filter(user=request.user).first()
    total_bookings = 0
    member_since   = getattr(request.user, 'created_at', request.user.date_joined)
    if customer:
        total_bookings = Bookings.objects.filter(customer=customer).count()

    context = {
        **get_customer_context(request),
        'active':         'profile',
        'customer':       customer,
        'total_bookings': total_bookings,
        'member_since':   member_since,
    }
    return render(request, 'garage/Customer/profile.html', context)


# ============================================================
#  CUSTOMER LOGOUT
# ============================================================
def customer_logout(request):
    auth_logout(request)
    messages.success(request, 'You have been logged out.')
    return redirect('login')


# ============================================================
#  SHARED PROVIDER CONTEXT
# ============================================================
def get_provider_context(request):
    provider      = ServiceProvider.objects.filter(user=request.user).first()
    pending_count = 0
    unread_count  = 0
    total_reviews = 0

    if provider:
        pending_count = Bookings.objects.filter(
            provider=provider, bookingStatus='pending'
        ).count()
        total_reviews = Review.objects.filter(provider=provider).count()

    unread_count = Notification.objects.filter(
        user=request.user, isRead=False
    ).count()

    return {
        'provider':               provider,
        'pending_bookings_count': pending_count,
        'unread_notif_count':     unread_count,
        'total_reviews':          total_reviews,
    }


# ============================================================
#  SERVICE PROVIDER REDIRECT
# ============================================================
@role_required(allowed_roles=["service_provider"])
def serviceProviderdashboard(request):
    return redirect('provider_overview')


# ============================================================
#  PROVIDER — OVERVIEW
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_overview(request):
    provider = ServiceProvider.objects.filter(user=request.user).first()
    if not provider:
        messages.warning(request, 'Your provider profile is not set up yet.')
        return redirect('provider_profile')

    today = date.today()

    total_bookings = Bookings.objects.filter(provider=provider).count()
    pending_count  = Bookings.objects.filter(provider=provider, bookingStatus='pending').count()
    monthly_revenue = (
        Payments.objects
        .filter(booking__provider=provider, paymentStatus='completed',
                paymentDate__month=today.month, paymentDate__year=today.year)
        .aggregate(total=Sum('amount'))['total'] or 0
    )
    monthly_count = Bookings.objects.filter(
        provider=provider,
        bookingDate__month=today.month,
        bookingDate__year=today.year
    ).count()

    completed       = Bookings.objects.filter(provider=provider, bookingStatus='completed').count()
    completion_rate = round((completed / total_bookings * 100) if total_bookings else 0)

    avg_duration = round(
        Services.objects.filter(providerId=provider)
        .aggregate(avg=AvgF('estimatedDuration'))['avg'] or 0
    )
    services_count = Services.objects.filter(providerId=provider, isAvailable=True).count()

    six_months_ago = (today.replace(day=1) - timedelta(days=180))
    monthly_qs = (
        Bookings.objects
        .filter(provider=provider, bookingDate__gte=six_months_ago)
        .annotate(month=TruncMonth('bookingDate'))
        .values('month')
        .annotate(count=Count('bookingId'))
        .order_by('month')
    )
    max_count = max((m['count'] for m in monthly_qs), default=1)
    monthly_bookings = [
        {
            'month': month_abbr[m['month'].month],
            'count': m['count'],
            'pct':   round(m['count'] / max_count * 100),
        }
        for m in monthly_qs
    ]

    todays_bookings = (
        Bookings.objects
        .filter(provider=provider, bookingDate=today)
        .select_related('customer__user').prefetch_related('services')
        .order_by('bookingTime')
    )

    context = {
        **get_provider_context(request),
        'active':           'overview',
        'today':            today.strftime('%d %b %Y'),
        'total_bookings':   total_bookings,
        'pending_count':    pending_count,
        'monthly_revenue':  monthly_revenue,
        'monthly_count':    monthly_count,
        'completion_rate':  completion_rate,
        'avg_duration':     avg_duration,
        'services_count':   services_count,
        'monthly_bookings': monthly_bookings,
        'todays_bookings':  todays_bookings,
    }
    return render(request, 'garage/Provider/overview.html', context)


# ============================================================
#  PROVIDER — BOOKINGS
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_bookings(request):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    status   = request.GET.get('status', 'all')
    page     = request.GET.get('page', 1)

    qs = (
        Bookings.objects
        .filter(provider=provider)
        .select_related('customer__user').prefetch_related('services')
        .order_by('-bookingDate', '-bookingTime')
    )
    if status not in ('', 'all'):
        qs = qs.filter(bookingStatus=status)

    counts = Bookings.objects.filter(provider=provider).aggregate(
        pending    = Count('bookingId', filter=Q(bookingStatus='pending')),
        inprogress = Count('bookingId', filter=Q(bookingStatus='in_progress')),
        completed  = Count('bookingId', filter=Q(bookingStatus='completed')),
        cancelled  = Count('bookingId', filter=Q(bookingStatus='cancelled')),
    )

    booking_list = list(qs)
    booking_ids  = [b.bookingId for b in booking_list]
    invoice_map  = {}
    for inv in Invoice.objects.filter(booking__bookingId__in=booking_ids):
        invoice_map[inv.booking_id] = inv

    for b in booking_list:
        b.has_invoice = b.bookingId in invoice_map
        b.invoice_obj = invoice_map.get(b.bookingId, None)

    paginator     = Paginator(booking_list, 12)
    bookings_page = paginator.get_page(page)

    context = {
        **get_provider_context(request),
        'active':           'bookings',
        'bookings':         bookings_page,
        'status':           status,
        'pending_count':    counts['pending'],
        'inprogress_count': counts['inprogress'],
        'completed_count':  counts['completed'],
        'cancelled_count':  counts['cancelled'],
    }
    return render(request, 'garage/Provider/bookings.html', context)


def provider_booking_detail(request, booking_id):
    """
    Handles both AJAX (eye-button modal) and normal page requests.
    Auth is checked inline so AJAX callers get a JSON error instead of
    a redirect that the browser cannot follow and that causes 'Failed to fetch'.
    """
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    # ── Auth guard (AJAX-aware) ──────────────────────────────
    if not request.user.is_authenticated:
        if is_ajax:
            return JsonResponse({'error': 'Session expired. Please log in again.'}, status=401)
        return redirect('login')

    provider = ServiceProvider.objects.filter(user=request.user).first()
    if not provider:
        if is_ajax:
            return JsonResponse({'error': 'Provider profile not found.'}, status=403)
        return redirect('login')

    booking = get_object_or_404(
        Bookings.objects.select_related('customer__user').prefetch_related('services'),
        bookingId=booking_id,
        provider=provider,
    )

    # ✅ Return JSON for AJAX eye-button requests
    if is_ajax:
        # ── vehicle ──────────────────────────────────────────────
        vehicle_info = '—'
        try:
            if getattr(booking, 'vehicle', None) and booking.vehicle:
                vehicle_info = (
                    f'{booking.vehicle.vehicleNumber} — {booking.vehicle.vehicleModel}'
                )
            elif getattr(booking.customer, 'vehicleNumber', None):
                vehicle_info = booking.customer.vehicleNumber
        except Exception:
            vehicle_info = getattr(booking.customer, 'vehicleNumber', '—') or '—'

        # ── services / price ─────────────────────────────────────
        try:
            service_text = getattr(booking, 'service_names', None) or 'N/A'
        except Exception:
            service_text = 'N/A'

        try:
            price_text = str(getattr(booking, 'total_price', 0) or 0)
        except Exception:
            price_text = '0'

        # ── customer info ────────────────────────────────────────
        try:
            _u = booking.customer.user
            _first = (_u.first_name or '').strip()
            _last  = (_u.last_name  or '').strip()
            full_name = f'{_first} {_last}'.strip()
            customer_name = full_name if full_name else (_u.email or '—')
        except Exception:
            customer_name = '—'

        try:
            customer_email = booking.customer.user.email
        except Exception:
            customer_email = '—'

        # ── date / time ──────────────────────────────────────────
        try:
            date_str = booking.bookingDate.strftime('%d %b %Y') if booking.bookingDate else '—'
        except Exception:
            date_str = '—'

        try:
            time_str = str(booking.bookingTime) if booking.bookingTime else '—'
        except Exception:
            time_str = '—'

        # ── notes ────────────────────────────────────────────────
        try:
            notes_str = getattr(booking, 'notes', None) or '—'
        except Exception:
            notes_str = '—'

        return JsonResponse({
            'bookingId':      booking.bookingId,
            'customer':       customer_name,
            'email':          customer_email,
            'vehicle':        vehicle_info,
            'service':        service_text,
            'price':          price_text,
            'date':           date_str,
            'time':           time_str,
            'status':         booking.bookingStatus,
            'status_display': booking.get_bookingStatus_display(),
            'notes':          notes_str,
        })

    # Normal page render (direct URL visit)
    has_invoice = False
    invoice = None
    try:
        invoice = booking.invoice
        if invoice:
            has_invoice = True
    except Exception:
        pass

    context = {
        **get_provider_context(request),
        'active':      'bookings',
        'booking':     booking,
        'has_invoice': has_invoice,
        'invoice':     invoice,
    }
    return render(request, 'garage/Provider/booking_detail.html', context)


@role_required(allowed_roles=["service_provider"])
@require_POST
def provider_confirm_booking(request, booking_id):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    booking  = get_object_or_404(Bookings, bookingId=booking_id, provider=provider)
    if booking.bookingStatus != 'pending':
        messages.warning(request, f'Booking #{booking.bookingId} cannot be confirmed (current status: {booking.get_bookingStatus_display()}).')
        return redirect('provider_bookings')
    booking.bookingStatus = 'confirmed'
    booking.save()
    Notification.objects.create(
        user=booking.customer.user,
        notificationType='booking_confirmed',
        title='Booking Confirmed',
        message=(
            f'Your booking #{booking.bookingId} for {booking.service_names} '
            f'at {provider.garageName} has been confirmed.'
        ),
    )
    messages.success(request, f'Booking #{booking.bookingId} confirmed.')
    return redirect('provider_bookings')


@role_required(allowed_roles=["service_provider"])
@require_POST
def provider_start_booking(request, booking_id):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    booking  = get_object_or_404(Bookings, bookingId=booking_id, provider=provider)
    if booking.bookingStatus != 'confirmed':
        messages.warning(request, f'Booking #{booking.bookingId} cannot be started (current status: {booking.get_bookingStatus_display()}).')
        return redirect('provider_bookings')
    booking.bookingStatus = 'in_progress'
    booking.save()
    Notification.objects.create(
        user=booking.customer.user,
        notificationType='general',
        title='Service Started',
        message=(
            f'Your {booking.service_names} service at '
            f'{provider.garageName} has started.'
        ),
    )
    messages.success(request, f'Booking #{booking.bookingId} is now in progress.')
    return redirect('provider_bookings')


@role_required(allowed_roles=["service_provider"])
@require_POST
def provider_complete_booking(request, booking_id):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    booking  = get_object_or_404(Bookings, bookingId=booking_id, provider=provider)
    if booking.bookingStatus != 'in_progress':
        messages.warning(request, f'Booking #{booking.bookingId} cannot be completed (current status: {booking.get_bookingStatus_display()}).')
        return redirect('provider_bookings')
    booking.bookingStatus = 'completed'
    booking.save()
    Notification.objects.create(
        user=booking.customer.user,
        notificationType='service_completed',
        title='Service Completed',
        message=(
            f'Your {booking.service_names} service by {provider.garageName} '
            f'is complete. Please leave a review!'
        ),
    )
    messages.success(request, f'Booking #{booking.bookingId} marked as completed. Generate invoice now.')
    return redirect('provider_invoice_generate', booking_id=booking.bookingId)


# ============================================================
#  PROVIDER — SERVICES
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_services(request):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    services_list = Services.objects.filter(providerId=provider).order_by('serviceName')
    context = {
        **get_provider_context(request),
        'active':   'services',
        'services': services_list,
    }
    return render(request, 'garage/Provider/services.html', context)


@role_required(allowed_roles=["service_provider"])
@require_POST
def provider_service_save(request):
    provider    = get_object_or_404(ServiceProvider, user=request.user)
    service_id   = request.POST.get('service_id', '').strip()
    name         = request.POST.get('name', '').strip()
    description  = request.POST.get('description', '').strip()
    price        = request.POST.get('price', 0)
    duration     = request.POST.get('duration') or None
    available    = request.POST.get('available', 'true') == 'true'
    vehicle_type = request.POST.get('vehicle_type', 'all').strip()  # ✅ FIX: was missing

    if not name:
        messages.error(request, 'Service name is required.')
        return redirect('provider_services')

    if service_id:
        svc = get_object_or_404(Services, pk=service_id, providerId=provider)
        svc.serviceName        = name
        svc.serviceDescription = description
        svc.servicePrice       = price
        svc.estimatedDuration  = duration
        svc.isAvailable        = available
        svc.vehicleType        = vehicle_type  # ✅ FIX: was never saved on edit
        svc.save()
        messages.success(request, f'"{name}" updated successfully.')
    else:
        Services.objects.create(
            providerId         = provider,
            serviceName        = name,
            serviceDescription = description,
            servicePrice       = price,
            estimatedDuration  = duration,
            isAvailable        = available,
            vehicleType        = vehicle_type,  # ✅ FIX: was never saved on create
        )
        messages.success(request, f'"{name}" added successfully.')

    return redirect('provider_services')


@role_required(allowed_roles=["service_provider"])
@require_POST
def provider_service_delete(request, pk):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    svc      = get_object_or_404(Services, pk=pk, providerId=provider)
    name     = svc.serviceName
    svc.delete()
    messages.success(request, f'"{name}" deleted.')
    return redirect('provider_services')


# ============================================================
#  PROVIDER — REVIEWS
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_reviews(request):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    reviews_qs = (
        Review.objects
        .filter(provider=provider)
        .select_related('customer__user')
        .order_by('-createdAt')
    )

    total        = reviews_qs.count()
    avg_r        = reviews_qs.aggregate(avg=Avg('rating'))['avg'] or 0
    avg_rating   = round(avg_r, 1)
    positive     = reviews_qs.filter(rating__gte=4).count()
    positive_pct = round((positive / total * 100) if total else 0)
    flagged_count = reviews_qs.filter(rating__lte=2).count()

    paginator   = Paginator(reviews_qs, 10)
    review_page = paginator.get_page(request.GET.get('page', 1))

    context = {
        **get_provider_context(request),
        'active':        'reviews',
        'reviews':       review_page,
        'avg_rating':    avg_rating,
        'positive_pct':  positive_pct,
        'flagged_count': flagged_count,
    }
    return render(request, 'garage/Provider/reviews.html', context)


# ============================================================
#  PROVIDER — EARNINGS
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_earnings(request):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    today    = date.today()

    payments_qs = (
        Payments.objects
        .filter(booking__provider=provider)
        .select_related('booking__customer__user')
        .order_by('-paymentDate')
    )
    totals = payments_qs.aggregate(
        total_earned   = Sum('amount', filter=Q(paymentStatus='completed')),
        pending_amount = Sum('amount', filter=Q(paymentStatus='pending')),
    )
    monthly_revenue = (
        payments_qs
        .filter(paymentStatus='completed',
                paymentDate__month=today.month,
                paymentDate__year=today.year)
        .aggregate(total=Sum('amount'))['total'] or 0
    )

    invoices_qs = (
        Invoice.objects
        .filter(booking__provider=provider)
        .select_related('booking__customer__user', 'payment')
        .order_by('-invoiceDate')
    )
    total_invoices = invoices_qs.count()

    pay_paginator = Paginator(payments_qs, 10)
    inv_paginator = Paginator(invoices_qs, 10)

    context = {
        **get_provider_context(request),
        'active':          'earnings',
        'payments':        pay_paginator.get_page(request.GET.get('pay_page', 1)),
        'invoices':        inv_paginator.get_page(request.GET.get('inv_page', 1)),
        'total_earned':    totals['total_earned']   or 0,
        'pending_amount':  totals['pending_amount'] or 0,
        'monthly_revenue': monthly_revenue,
        'total_invoices':  total_invoices,
    }
    return render(request, 'garage/Provider/earnings.html', context)


# ============================================================
# provider payout request, payout history, etc. can be implemented here
# ============================================================
@login_required
@require_POST
def provider_complete_payment(request):
    try:
        data = _json.loads(request.body)
        pay_id = data.get('payment_id')

        payment = Payments.objects.get(
            paymentId=pay_id,
            paymentStatus='pending'
        )
        payment.paymentStatus = 'completed'
        payment.save()
        return JsonResponse({'success': True})

    except Payments.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Payment not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})






# ============================================================
#  PROVIDER — NOTIFICATIONS
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_notifications(request):
    notifs = Notification.objects.filter(user=request.user).order_by('-createdAt')
    total  = notifs.count()
    read   = notifs.filter(isRead=True).count()
    unread = notifs.filter(isRead=False).count()
    open_rate = round((read / total * 100) if total else 0)

    paginator  = Paginator(notifs, 15)
    notif_page = paginator.get_page(request.GET.get('page', 1))

    context = {
        **get_provider_context(request),
        'active':        'notifications',
        'notifications': notif_page,
        'total_notifs':  total,
        'read_notifs':   read,
        'open_rate':     open_rate,
    }
    return render(request, 'garage/Provider/notifications.html', context)


@role_required(allowed_roles=["service_provider"])
def provider_notification_read(request, pk):
    notif = get_object_or_404(Notification, pk=pk, user=request.user)
    notif.isRead = True
    notif.save()
    return redirect('provider_notifications')


@role_required(allowed_roles=["service_provider"])
def provider_notifications_read_all(request):
    Notification.objects.filter(user=request.user, isRead=False).update(isRead=True)
    messages.success(request, 'All notifications marked as read.')
    return redirect('provider_notifications')


# ============================================================
#  PROVIDER — PROFILE
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_profile(request):
    provider = ServiceProvider.objects.filter(user=request.user).first()

    if request.method == 'POST':
        action = request.POST.get('action', 'garage')

        if action == 'garage':
            if not provider:
                provider = ServiceProvider(user=request.user)
            provider.garageName  = request.POST.get('garage_name', '').strip()
            provider.description = request.POST.get('description', '').strip()
            provider.location    = request.POST.get('location', '').strip()
            lat = request.POST.get('latitude', '')
            lng = request.POST.get('longitude', '')
            provider.latitude  = float(lat) if lat else None
            provider.longitude = float(lng) if lng else None
            ot = request.POST.get('opening_time')
            ct = request.POST.get('closing_time')
            if ot: provider.openingTime = ot
            if ct: provider.closingTime = ct
            if 'garage_image' in request.FILES:
                provider.garageImage = request.FILES['garage_image']
            provider.save()
            messages.success(request, 'Garage info updated successfully!')

        elif action == 'account':
            request.user.first_name = request.POST.get('first_name', '').strip()
            request.user.last_name  = request.POST.get('last_name', '').strip()
            if hasattr(request.user, 'mobile'):
                request.user.mobile = request.POST.get('mobile', '').strip()
            request.user.save()
            messages.success(request, 'Account info updated!')

        elif action == 'password':
            cur  = request.POST.get('current_password', '')
            new  = request.POST.get('new_password', '')
            conf = request.POST.get('confirm_password', '')
            if not request.user.check_password(cur):
                messages.error(request, 'Current password is incorrect.')
            elif new != conf:
                messages.error(request, 'Passwords do not match.')
            elif len(new) < 8:
                messages.error(request, 'Password must be at least 8 characters.')
            else:
                request.user.set_password(new)
                request.user.save()
                messages.success(request, 'Password updated! Please log in again.')
                return redirect('login')

        return redirect('provider_profile')

    total_bookings = Bookings.objects.filter(provider=provider).count() if provider else 0
    services_count = Services.objects.filter(providerId=provider).count() if provider else 0

    context = {
        **get_provider_context(request),
        'active':         'profile',
        'user':           request.user,
        'total_bookings': total_bookings,
        'services_count': services_count,
    }
    return render(request, 'garage/Provider/profile.html', context)


# ============================================================
#  PROVIDER LOGOUT
# ============================================================
def provider_logout(request):
    auth_logout(request)
    messages.success(request, 'You have been logged out.')
    return redirect('login')


# ============================================================
#  INVOICE HELPER
# ============================================================
def _next_invoice_number(provider):
    """Generate sequential invoice number like INV-2026-001."""
    year  = date.today().year
    count = Invoice.objects.filter(
        booking__provider=provider,
        invoiceDate__year=year
    ).count() + 1
    return f"INV-{year}-{count:03d}"


# ============================================================
#  PROVIDER — GENERATE INVOICE
# ============================================================
@role_required(allowed_roles=["service_provider"])
@role_required(allowed_roles=["service_provider"])
def provider_invoice_generate(request, booking_id):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    booking  = get_object_or_404(
        Bookings.objects.select_related('customer__user').prefetch_related('services'),
        bookingId=booking_id,
        provider=provider,
    )
    if booking.bookingStatus != 'completed':
        messages.warning(request, f'Invoice can only be generated for completed bookings (current status: {booking.get_bookingStatus_display()}).')
        return redirect('provider_bookings')

    existing_invoice = None
    try:
        existing_invoice = booking.invoice
    except Exception:
        pass

    if existing_invoice:
        messages.info(request, 'Invoice already exists. Redirecting to view.')
        return redirect('provider_invoice_view', invoice_id=existing_invoice.invoiceId)

    if request.method == 'POST':
        inv_num      = request.POST.get('invoice_number', '').strip()
        gst_pct      = float(request.POST.get('gst_percent', 18))
        gst_amount   = float(request.POST.get('gst_amount') or 0)
        discount     = float(request.POST.get('discount_amount') or 0)
        total_amount = float(request.POST.get('total_amount', booking.total_price))
        pay_method   = request.POST.get('payment_method', 'cash')
        pay_status   = request.POST.get('payment_status', 'completed')
        notes        = request.POST.get('notes', '').strip()

        if not inv_num:
            inv_num = _next_invoice_number(provider)

        payment = Payments.objects.create(
            booking       = booking,
            amount        = total_amount,
            paymentMethod = pay_method,
            paymentStatus = pay_status,
            transactionId = str(uuid.uuid4())[:12].upper(),
            paymentDate   = timezone.now(),
        )

        invoice = Invoice.objects.create(
            booking        = booking,
            payment        = payment,
            invoiceNumber  = inv_num,
            invoiceDate    = date.today(),
            totalAmount    = total_amount,
            taxAmount      = gst_amount,
            discountAmount = discount,
        )

        Notification.objects.create(
            user             = booking.customer.user,
            notificationType = 'payment_received',
            title            = 'Invoice Generated',
            message          = (
                f'Invoice {inv_num} has been generated for your '
                f'{booking.service_names} booking. '
                f'Total: ₹{total_amount}.'
            ),
        )

        # ── Send email to customer ──────────────────────────────
        customer_email = booking.customer.user.email
        customer_name  = booking.customer.user.get_full_name() or customer_email
        invoice_url    = request.build_absolute_uri(
            f'/garage/serviceProvider/invoice/{invoice.invoiceId}/view/'
        )

        email_subject = f'Your Service is Completed – Invoice {inv_num} | {provider.garageName}'

        email_body = f"""Dear {customer_name},

Your vehicle service at {provider.garageName} has been successfully completed.

──────────────────────────────
  Invoice Details
──────────────────────────────
  Invoice No  : {inv_num}
  Services    : {booking.service_names}
  Date        : {date.today().strftime('%d %b %Y')}
  Total Amount: ₹{total_amount:.2f}
  Payment     : {pay_method.upper()} — {pay_status.capitalize()}
──────────────────────────────

You can view and download your invoice here:
{invoice_url}

Thank you for choosing {provider.garageName}!
We look forward to serving you again.

Regards,
{provider.garageName}
eGarage Team
"""

        try:
            send_mail(
                subject      = email_subject,
                message      = email_body,
                from_email   = settings.DEFAULT_FROM_EMAIL,
                recipient_list = [customer_email],
                fail_silently  = False,
            )
        except Exception as mail_err:
            # Don't block invoice creation if email fails — just log it
            print(f'[eGarage] Invoice email failed for {customer_email}: {mail_err}')

        messages.success(request, f'Invoice {inv_num} generated & email sent to {customer_email}!')
        return redirect('provider_invoice_view', invoice_id=invoice.invoiceId)

    context = {
        **get_provider_context(request),
        'active':           'earnings',
        'booking':          booking,
        'user':             request.user,
        'today':            date.today().strftime('%d %b %Y'),
        'next_invoice_num': _next_invoice_number(provider).split('-')[-1],
    }
    return render(request, 'garage/Provider/invoice_generate.html', context)


# ============================================================
#  PROVIDER — VIEW INVOICE
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_invoice_view(request, invoice_id):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    invoice  = get_object_or_404(
        Invoice.objects.select_related(
            'booking__customer__user', 'payment'
        ),
        invoiceId=invoice_id,
        booking__provider=provider,
    )
    context = {
        **get_provider_context(request),
        'active':   'earnings',
        'invoice':  invoice,
        'provider': provider,
        'user':     request.user,
        'today':    date.today().strftime('%d %b %Y'),
    }
    return render(request, 'garage/Provider/invoice_view.html', context)


# ============================================================
#  PROVIDER — DOWNLOAD INVOICE CSV
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_invoice_download(request, invoice_id):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    invoice  = get_object_or_404(
        Invoice.objects.select_related(
            'booking__customer__user', 'payment'
        ),
        invoiceId=invoice_id,
        booking__provider=provider,
    )
    b = invoice.booking

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="' + invoice.invoiceNumber + '.csv"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    writer = csv.writer(response)
    writer.writerow(['E-GARAGE — TAX INVOICE'])
    writer.writerow([])
    writer.writerow(['Invoice No.',    invoice.invoiceNumber])
    writer.writerow(['Invoice Date',   str(invoice.invoiceDate)])
    writer.writerow(['Booking ID',     '#BK' + str(b.bookingId)])
    writer.writerow([])
    writer.writerow(['FROM (Provider)', provider.garageName])
    writer.writerow(['Location',        provider.location])
    writer.writerow(['Email',           request.user.email])
    writer.writerow([])
    writer.writerow(['TO (Customer)',   b.customer.user.get_full_name()])
    writer.writerow(['Email',           b.customer.user.email])
    writer.writerow(['Vehicle',         b.customer.vehicleNumber])
    writer.writerow([])
    writer.writerow(['DESCRIPTION',     'AMOUNT'])
    writer.writerow([b.service_names, 'Rs.' + str(b.total_price)])
    writer.writerow(['GST',             'Rs.' + str(invoice.taxAmount)])
    writer.writerow(['Discount',        'Rs.' + str(invoice.discountAmount)])
    writer.writerow(['TOTAL',           'Rs.' + str(invoice.totalAmount)])
    if invoice.payment:
        writer.writerow([])
        writer.writerow(['Payment Method', invoice.payment.paymentMethod])
        writer.writerow(['Payment Status', invoice.payment.paymentStatus])
        writer.writerow(['Transaction ID', invoice.payment.transactionId or '—'])
    return response


# ============================================================
#  PROVIDER — PRINT INVOICE
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_invoice_print(request, invoice_id):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    invoice  = get_object_or_404(
        Invoice.objects.select_related(
            'booking__customer__user', 'payment'
        ),
        invoiceId=invoice_id,
        booking__provider=provider,
    )
    b = invoice.booking
    base_amt = float(invoice.totalAmount) - float(invoice.taxAmount) + float(invoice.discountAmount)

    pay_method = '—'
    pay_status = '—'
    pay_txn    = '—'
    if invoice.payment:
        pay_method = invoice.payment.paymentMethod or '—'
        pay_status = invoice.payment.paymentStatus or '—'
        pay_txn    = invoice.payment.transactionId or '—'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Invoice {invoice.invoiceNumber}</title>
<style>
@page {{ size: A4; margin: 15mm; }}
* {{ margin:0; padding:0; box-sizing:border-box; font-family:'Segoe UI',Arial,sans-serif; }}
body {{ padding:30px; color:#222; background:#fff; }}
.inv-box {{ max-width:750px; margin:auto; }}
.top-bar {{ background:#e8560a; color:#fff; padding:18px 28px; display:flex;
            justify-content:space-between; align-items:center; border-radius:8px 8px 0 0; }}
.top-bar h1 {{ font-size:22px; }}
.top-bar .inv-label {{ text-align:right; }}
.top-bar .inv-label h2 {{ font-size:16px; font-weight:400; opacity:.85; }}
.top-bar .inv-label p {{ font-size:20px; font-weight:700; }}
.body-area {{ border:1px solid #ddd; border-top:none; padding:28px; border-radius:0 0 8px 8px; }}
.parties {{ display:flex; justify-content:space-between; margin-bottom:24px; }}
.parties div {{ width:48%; }}
.parties h4 {{ color:#e8560a; font-size:12px; text-transform:uppercase; letter-spacing:.5px;
               margin-bottom:6px; }}
.parties p {{ font-size:14px; line-height:1.7; color:#444; }}
.meta-row {{ display:flex; justify-content:space-between; background:#f8f8f8;
             padding:10px 16px; border-radius:6px; margin-bottom:20px; font-size:13px; color:#555; }}
table {{ width:100%; border-collapse:collapse; margin-bottom:18px; }}
thead {{ background:#1e3a5f; color:#fff; }}
thead th {{ padding:10px 14px; text-align:left; font-size:13px; font-weight:600; }}
tbody td {{ padding:10px 14px; border-bottom:1px solid #eee; font-size:14px; }}
.total-row td {{ font-weight:700; font-size:16px; border-top:2px solid #1e3a5f; background:#f0f7ff; }}
.pay-info {{ background:#f8f8f8; padding:14px 18px; border-radius:6px; font-size:13px; color:#555;
             display:flex; gap:30px; margin-bottom:18px; }}
.pay-info span {{ font-weight:600; color:#222; }}
.footer {{ text-align:center; padding-top:16px; border-top:1px solid #eee; }}
.footer p {{ font-size:12px; color:#999; }}
.footer p.thanks {{ font-size:14px; color:#e8560a; font-weight:600; margin-bottom:4px; }}
.print-btn {{ text-align:center; margin-bottom:18px; }}
.print-btn button {{ background:#e8560a; color:#fff; border:none; padding:10px 32px;
                     font-size:14px; font-weight:600; border-radius:6px; cursor:pointer; }}
.print-btn button:hover {{ background:#d14a00; }}
@media print {{ .print-btn {{ display:none; }} }}
</style>
</head>
<body>
<div class="print-btn">
  <button onclick="window.print()">🖨️ Print Invoice</button>
  <button onclick="window.close()" style="background:#6b7280;margin-left:8px">✕ Close</button>
</div>
<div class="inv-box">
  <div class="top-bar">
    <div><h1>e<span style="color:#ffd699">Garage</span></h1><p style="font-size:12px;opacity:.8">TAX INVOICE</p></div>
    <div class="inv-label"><h2>Invoice</h2><p>{invoice.invoiceNumber}</p></div>
  </div>
  <div class="body-area">
    <div class="parties">
      <div>
        <h4>From (Service Provider)</h4>
        <p><strong>{provider.garageName}</strong><br>
        {provider.location}<br>
        {request.user.email}<br>
        {getattr(request.user, 'mobile', '') or ''}</p>
      </div>
      <div>
        <h4>To (Customer)</h4>
        <p><strong>{b.customer.user.get_full_name()}</strong><br>
        {b.customer.user.email}<br>
        Vehicle: {b.customer.vehicleNumber}<br>
        {b.customer.vehicleModel or ''} {b.customer.vehicleColor or ''}</p>
      </div>
    </div>
    <div class="meta-row">
      <span>Invoice Date: <strong>{invoice.invoiceDate}</strong></span>
      <span>Booking: <strong>#BK{b.bookingId}</strong></span>
      <span>Service Date: <strong>{b.bookingDate}</strong></span>
    </div>
    <table>
      <thead><tr><th>Description</th><th style="text-align:right">Amount</th></tr></thead>
      <tbody>
        <tr><td>{b.service_names}</td><td style="text-align:right">₹{base_amt:.2f}</td></tr>
        <tr><td>GST / Tax</td><td style="text-align:right">₹{invoice.taxAmount:.2f}</td></tr>
        <tr><td>Discount</td><td style="text-align:right">- ₹{invoice.discountAmount:.2f}</td></tr>
        <tr class="total-row"><td>Total Payable</td><td style="text-align:right">₹{invoice.totalAmount:.2f}</td></tr>
      </tbody>
    </table>
    <div class="pay-info">
      <div>Method: <span>{pay_method}</span></div>
      <div>Status: <span>{pay_status}</span></div>
      <div>Txn ID: <span>{pay_txn}</span></div>
    </div>
    <div class="footer">
      <p class="thanks">Thank you for choosing {provider.garageName}!</p>
      <p>This is a computer-generated invoice. No signature required.</p>
      <p>eGarage &bull; Generated on {date.today().strftime('%d %b %Y')}</p>
    </div>
  </div>
</div>
<script>setTimeout(function(){{ window.print(); }}, 600);</script>
</body>
</html>"""
    return HttpResponse(html, content_type='text/html; charset=utf-8')


# ============================================================
#  PROVIDER — EXPORT ALL INVOICES CSV
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_invoices_export(request):
    provider = get_object_or_404(ServiceProvider, user=request.user)
    invoices_qs = Invoice.objects.filter(
        booking__provider=provider
    ).select_related('booking__customer__user', 'payment').order_by('-invoiceDate')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="{provider.garageName}_invoices.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(['Invoice No', 'Date', 'Booking', 'Customer', 'Service', 'Vehicle',
                     'Service Amount', 'GST', 'Discount', 'Total', 'Payment Status'])
    for inv in invoices_qs:
        b = inv.booking
        writer.writerow([
            inv.invoiceNumber,
            inv.invoiceDate,
            f'#BK{b.bookingId}',
            b.customer.user.get_full_name(),
            b.service_names,
            b.customer.vehicleNumber,
            b.total_price,
            inv.taxAmount,
            inv.discountAmount,
            inv.totalAmount,
            inv.payment.paymentStatus if inv.payment else 'No Payment',
        ])
    return response


# ============================================================
#  PROVIDER — EXPORT EARNINGS CSV
# ============================================================
@role_required(allowed_roles=["service_provider"])
def provider_earnings_export(request):
    provider = get_object_or_404(ServiceProvider, user=request.user)

    filename = provider.garageName.replace(' ', '_') + '_earnings_' + str(date.today()) + '.csv'
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="' + filename + '"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    writer = csv.writer(response)

    writer.writerow(['=== PAYMENT HISTORY ==='])
    writer.writerow([
        'Pay ID', 'Booking', 'Customer', 'Service',
        'Amount', 'Method', 'Date', 'Status',
    ])
    pay_qs = (
        Payments.objects
        .filter(booking__provider=provider)
        .select_related('booking__customer__user')
        .order_by('-paymentDate')
    )
    for pay in pay_qs:
        writer.writerow([
            '#PAY' + str(pay.paymentId),
            '#BK' + str(pay.booking.bookingId),
            pay.booking.customer.user.get_full_name(),
            pay.booking.service_names,
            str(pay.amount),
            pay.get_paymentMethod_display(),
            pay.paymentDate.strftime('%d %b %Y') if pay.paymentDate else '',
            pay.get_paymentStatus_display(),
        ])

    writer.writerow([])
    writer.writerow(['=== INVOICE HISTORY ==='])
    writer.writerow([
        'Invoice No', 'Booking', 'Customer', 'Service', 'Vehicle',
        'GST', 'Discount', 'Total', 'Date', 'Status',
    ])
    inv_qs = (
        Invoice.objects
        .filter(booking__provider=provider)
        .select_related('booking__customer__user', 'payment')
        .order_by('-invoiceDate')
    )
    for inv in inv_qs:
        b = inv.booking
        writer.writerow([
            inv.invoiceNumber,
            '#BK' + str(b.bookingId),
            b.customer.user.get_full_name(),
            b.service_names,
            b.customer.vehicleNumber,
            str(inv.taxAmount),
            str(inv.discountAmount),
            str(inv.totalAmount),
            inv.invoiceDate.strftime('%d %b %Y') if inv.invoiceDate else '',
            inv.payment.get_paymentStatus_display() if inv.payment else 'Unpaid',
        ])

    return response